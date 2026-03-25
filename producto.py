from abc import ABC , abstractmethod
from openpyxl import load_workbook


class Producto: 

    def __init__(self, id_producto : str, nombre : str, 
        descripcion : str, precio : float, stock : int):

        self._id_producto = id_producto
        self._nombre = nombre
        self._descripcion = descripcion
        self._precio = precio
        self._stock = stock

    @property
    def id_producto(self):
        return self._id_producto
    
    @property
    def nombre(self):
        return self._nombre
    
    @property
    def precio(self):
        return self._precio
    
    @property
    def stock(self):
        return self._stock

    @stock.setter
    def stock(self, nuevo_stock: int):
        self._stock = nuevo_stock


class Catalogo:
    def __init__(self):
        self._productos = []

    @property
    def productos(self):
        return self._productos
    
    # Cargar productos al catálogo 
    def agregar_producto(self, producto : Producto) -> None:
        self._productos.append(producto)

    def mostrar_catalogo(self) -> None:
        for producto in self.productos:
            print(f"ID: {producto.id_producto} - "
                  f"{producto.nombre} - ${producto.precio} - "
                  f"Stock: {producto.stock}")
    
    def buscar_producto(self, id_producto : str):
        for producto in self._productos:
            if producto.id_producto == id_producto:
                return producto


class Suscripcion:

    def __init__(self):
        self._status = "Activa"

    @property
    def status(self):
        return self._status
    
    def cancelar(self):
        self._status = "Cancelada"
    
    def reactivar(self):
        self._status = "Activa"


# Clase abstracta para inventarios - depende del sistema de inventario externo
class Inventario(ABC):

    @abstractmethod
    def cargar_inventario_externo(self):
        pass

    @abstractmethod
    def modificar_inventario_externo(self, producto : Producto, 
                                     cantidad : int):
        pass    


# Implementamos de forma supuesta para el ejercicio  un 
# sistema de inventario externo basado en excel
class InventarioExcel(Inventario):
    def __init__(self, ruta_archivo : str):
        self._ruta_archivo = ruta_archivo

    def cargar_inventario_externo(self) -> list:
        wb = load_workbook(self._ruta_archivo)
        ws = wb.active
        productos = []
        for row in ws.iter_rows(min_row=2, values_only=True): # type: ignore
            id_producto, nombre, descripcion, precio, stock = row
            producto = Producto(id_producto, nombre, # type: ignore
                                descripcion, precio, stock) # type: ignore
            productos.append(producto)
        wb.close()
        return productos
    
    def modificar_inventario_externo(self, producto : Producto, 
                                     cantidad : int) -> None:
        wb = load_workbook(self._ruta_archivo)
        ws = wb.active
        for row in ws.iter_rows(min_row=2): # type: ignore
            if row[0].value == producto.id_producto:
                row[4].value = cantidad # type: ignore
                break
        wb.save(self._ruta_archivo)
        wb.close()
        producto.stock = cantidad